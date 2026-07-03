"""Shared Excel import/export helpers (openpyxl)."""

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def workbook_response(filename, sheet_title, headers, rows):
    """Build an RTL worksheet with a bold header row and return it as a download."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(header)) + 6)
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def read_sheet(file_obj):
    """Read the first worksheet: returns (headers, data_rows) with trimmed strings."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else '' for h in next(rows)]
    except StopIteration:
        return [], []
    data = []
    for row in rows:
        if row is None or all(v in (None, '') for v in row):
            continue
        data.append([str(v).strip() if isinstance(v, str) else v for v in row])
    return headers, data
